import os
import csv

class ExportService:
    @staticmethod
    def export_to_csv(data, filename, headers=None):

        try:
            os.makedirs(os.path.dirname(os.path.abspath(filename)), exist_ok=True)

            if data and isinstance(data[0], dict):
                if headers is None:
                    headers = list(data[0].keys())
                rows = [list(item.values()) for item in data]
            else:
                rows = data

            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                if headers:
                    writer.writerow(headers)
                writer.writerows(rows)

            return {
                'success': True,
                'filename': filename,
                'rows_count': len(rows)
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }

    @staticmethod
    def export_to_excel(data, filename, sheet_name='Sheet1', headers=None):
        try:
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill
            except ImportError:
                return {
                    'success': False,
                    'error': 'Для экспорта в Excel требуется библиотека openpyxl. Установите её с помощью pip install openpyxl'
                }

            os.makedirs(os.path.dirname(os.path.abspath(filename)), exist_ok=True)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            if data and isinstance(data[0], dict):
                if headers is None:
                    headers = list(data[0].keys())
                rows = [list(item.values()) for item in data]
            else:
                rows = data

            if headers:
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            for row_idx, row_data in enumerate(rows, 2 if headers else 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filename)

            return {
                'success': True,
                'filename': filename,
                'rows_count': len(rows)
            }

        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }

    @staticmethod
    def format_report_data(report_data):
        report_type = report_data.get('report_name', '').lower()

        if 'просроченным задачам' in report_type:
            headers = ['ID', 'Описание', 'Статус', 'Часы', 'Проект', 'Дедлайн', 'Разработчик']
            rows = []
            for task in report_data.get('tasks', []):
                rows.append([
                    task.get('id', ''),
                    task.get('description', ''),
                    task.get('status', ''),
                    task.get('hours_worked', ''),
                    task.get('project_name', ''),
                    task.get('project_deadline', ''),
                    task.get('developer_name', '')
                ])

        elif 'загрузке разработчиков' in report_type:
            headers = ['ID', 'ФИО', 'Должность', 'Ставка', 'Кол-во задач', 'Часы', 'Стоимость']
            rows = []
            for dev in report_data.get('developers', []):
                rows.append([
                    dev.get('id', ''),
                    dev.get('full_name', ''),
                    dev.get('position', ''),
                    dev.get('hourly_rate', ''),
                    dev.get('task_count', ''),
                    dev.get('total_hours', ''),
                    dev.get('total_cost', '')
                ])

        elif 'статусу проектов' in report_type:
            headers = ['ID', 'Название', 'Клиент', 'Дедлайн', 'Бюджет', 'Задачи', 'Завершено', 'Прогресс', 'Часы',
                       'Стоимость']
            rows = []
            for proj in report_data.get('projects', []):
                rows.append([
                    proj.get('id', ''),
                    proj.get('name', ''),
                    proj.get('client', ''),
                    proj.get('deadline', ''),
                    proj.get('budget', ''),
                    proj.get('total_tasks', ''),
                    proj.get('completed_tasks', ''),
                    f"{proj.get('progress_percent', '')}%",
                    proj.get('total_hours', ''),
                    proj.get('total_cost', '')
                ])

        elif 'доходам за месяц' in report_type:
            headers = ['ID', 'Название', 'Клиент', 'Бюджет', 'Задачи', 'Часы', 'Стоимость', 'Прибыль']
            rows = []
            for proj in report_data.get('projects', []):
                rows.append([
                    proj.get('id', ''),
                    proj.get('name', ''),
                    proj.get('client', ''),
                    proj.get('budget', ''),
                    proj.get('task_count', ''),
                    proj.get('total_hours', ''),
                    proj.get('total_cost', ''),
                    proj.get('profit', '')
                ])

        else:
            if isinstance(report_data.get('data', []), list) and len(report_data.get('data', [])) > 0:
                first_item = report_data['data'][0]
                headers = list(first_item.keys())
                rows = [list(item.values()) for item in report_data['data']]
            else:
                headers = ['Данные']
                rows = [['Нет данных для экспорта']]

        return headers, rows
