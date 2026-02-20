import sys
import os
import unittest
import tempfile
import csv
import json

# Добавляем родительскую директорию в путь для импорта
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from controllers.export_controller import ExportController
from services.report_service import ReportService

class TestExport(unittest.TestCase):
    """
    Тесты для функционала экспорта
    """
    
    def setUp(self):
        """
        Настройка перед каждым тестом
        """
        self.export_controller = ExportController()
        self.report_service = ReportService()
        
        # Создаем временную директорию для тестовых файлов
        self.temp_dir = tempfile.mkdtemp()
    
    def tearDown(self):
        """
        Очистка после каждого теста
        """
        # Удаляем временные файлы
        for file in os.listdir(self.temp_dir):
            os.remove(os.path.join(self.temp_dir, file))
        os.rmdir(self.temp_dir)
    
    def test_export_to_csv(self):
        """
        Тест экспорта в CSV
        """
        # Подготовка тестовых данных
        data = [
            {'id': 1, 'name': 'Тест 1', 'value': 100},
            {'id': 2, 'name': 'Тест 2', 'value': 200},
            {'id': 3, 'name': 'Тест 3', 'value': 300}
        ]
        headers = ['id', 'name', 'value']
        
        # Экспорт в CSV
        filename = os.path.join(self.temp_dir, 'test_export.csv')
        result = self.export_controller.export_data_to_csv(data, headers, filename)
        
        # Проверка результата
        self.assertTrue(result['success'])
        self.assertTrue(os.path.exists(filename))
        
        # Проверка содержимого файла
        with open(filename, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            rows = list(reader)
            
            # Проверка заголовков
            self.assertEqual(rows[0], headers)
            
            # Проверка данных
            self.assertEqual(len(rows), 4)  # 1 строка заголовков + 3 строки данных
            self.assertEqual(rows[1][0], '1')
            self.assertEqual(rows[1][1], 'Тест 1')
            self.assertEqual(rows[1][2], '100')
    
    def test_export_report_to_csv(self):
        """
        Тест экспорта отчета в CSV
        """
        # Подготовка тестовых данных отчета
        report_data = {
            'report_name': 'Тестовый отчет',
            'generated_at': '2023-01-01 12:00:00',
            'projects': [
                {'id': 1, 'name': 'Проект 1', 'client': 'Клиент 1', 'budget': 100000},
                {'id': 2, 'name': 'Проект 2', 'client': 'Клиент 2', 'budget': 200000}
            ]
        }
        
        # Экспорт отчета в CSV
        filename = os.path.join(self.temp_dir, 'test_report.csv')
        result = self.export_controller.export_report_to_csv(report_data, filename)
        
        # Проверка результата
        self.assertTrue(result['success'])
        self.assertTrue(os.path.exists(filename))
    
    def test_export_to_excel(self):
        """
        Тест экспорта в Excel
        """
        try:
            import openpyxl
        except ImportError:
            self.skipTest("Библиотека openpyxl не установлена")
        
        # Подготовка тестовых данных
        data = [
            {'id': 1, 'name': 'Тест 1', 'value': 100},
            {'id': 2, 'name': 'Тест 2', 'value': 200},
            {'id': 3, 'name': 'Тест 3', 'value': 300}
        ]
        headers = ['id', 'name', 'value']
        
        # Экспорт в Excel
        filename = os.path.join(self.temp_dir, 'test_export.xlsx')
        result = self.export_controller.export_data_to_excel(data, headers, filename)
        
        # Проверка результата
        self.assertTrue(result['success'])
        self.assertTrue(os.path.exists(filename))
        
        # Проверка содержимого файла
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        
        # Проверка заголовков
        self.assertEqual(ws.cell(1, 1).value, 'id')
        self.assertEqual(ws.cell(1, 2).value, 'name')
        self.assertEqual(ws.cell(1, 3).value, 'value')
        
        # Проверка данных
        self.assertEqual(ws.cell(2, 1).value, 1)
        self.assertEqual(ws.cell(2, 2).value, 'Тест 1')
        self.assertEqual(ws.cell(2, 3).value, 100)

if __name__ == '__main__':
    unittest.main()
